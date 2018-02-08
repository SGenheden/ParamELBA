# Author: Samuel Genheden samuel.genheden@gmail.com

"""
Program to plot bond and angle distributions from a trajectory

The trajectory is assumed to be created by md_bondang.py

The force field definition, i.e. the bond and angles to be tracked
are stored in an XML format and read by the classes in the elbalib
module.

A reference structure need to be given and this is used to find CG
molecules defined in the XML file.

Examples
--------
anal_bondang.py -f bondang_elba.txt bondang_slipid.txt  -l elba slipid -r popc_elba.pdb -x elba.xml
"""

import argparse
import sys

import matplotlib.pylab as plt
import MDAnalysis as md
import numpy as np
import scipy.stats as stats
import scipy.optimize as opt
import openpyxl as xl

import elbalib
from sgenlib import colors


class TimeSeries(object):
    """
    Simple class to hold a time series and processes thereof
    """

    def __init__(self, data, transf=None):
        self.data0 = data
        if transf is None:
            self.data = data
        else:
            self.data = transf(data)
        self.len = self.data.shape[0]
        self.mean = self.data.mean()
        self.std = self.data.std()
        self.var = self.data.var()
#        self.kde = stats.gaussian_kde(self.data)
#        self.kde.covariance_factor = lambda: 0.25
#        self.kde._compute_covariance()
        self.histogram, self.edges = np.histogram(self.data, bins=self.len / 40.0)
        self.points = np.arange(self.mean-3.0*self.std,self.mean+3.5*self.std,0.5*self.std)

    def density(self, points=None):
        """
        Return the density of the Kernal density estimator
        """
        if points is None:
#            return self.kde(self.edges)
            return  stats.norm.pdf(self.points,loc=self.mean,scale=self.std)
        else:
#            return self.kde(points)
            return  stats.norm.pdf(points,loc=self.mean,scale=self.std)

    def histo(self):
        h,e = np.histogram(self.data,bins=self.points)
        return h/float(self.data.shape[0])

    def shapiro(self):
        return stats.shapiro(self.data)[0]


def _fit_ff(serieslist, labels, fflist, worksheet,offset=0):
    """
    Fit the distributions and compute
    """

    # This is the function we will be fitting to
    def _harmonic(xdata, k, eq, C):
        return k * (xdata - eq) * (xdata - eq) + C

    nseries = len(fflist)
    RT = 303.0 * 8.314 / 4.184 / 1000.0

    if isinstance(fflist[0].contype, elbalib.BondType) :
        rowoffset = offset
        ws.cell(row=1,column=1).value = "Bond"
    else:
        rowoffset = offset+4
        ws.cell(row=rowoffset+1,column=1).value = "Angle"
    labels2 = ["Ff"]
    labels2.extend(labels)
    for i,label in enumerate(labels2):
        ws.cell(row=rowoffset+1,column=1+i*3).value = label
        ws.cell(row=rowoffset+2,column=1+i*3).value = "k"
        if rowoffset == 0:
            ws.cell(row=rowoffset+2,column=1+i*3+1).value = "r_eq"
        else:
            ws.cell(row=rowoffset+2,column=1+i*3+1).value = "theta_eq"

    for iseries in range(nseries):
        # Print out the force field parameters
        fftype = fflist[iseries].contype
        isbond = isinstance(fftype, elbalib.BondType)
        ws.cell(row=rowoffset+3+iseries,column=1).value = fflist[iseries].atomstring()
        if isbond:
            print "%8s | ff: k=%9.3f r_eq=%9.3f" % (fflist[iseries].atomstring(), fftype.k * fftype.kf, fftype.r0),
            ws.cell(row=rowoffset+3+iseries,column=3).value = fftype.r0 * 0.1
            ws.cell(row=rowoffset+3+iseries,column=2).value = fftype.k * fftype.kf * 4.184 * 100
        else:
            print "%12s | ff: k=%9.3f theta_eq=%9.3f" % (fflist[iseries].atomstring(), fftype.k * fftype.kf, fftype.theta0),
            ws.cell(row=rowoffset+3+iseries,column=3).value = fftype.theta0
            ws.cell(row=rowoffset+3+iseries,column=2).value = fftype.k * fftype.kf * 4.184
        # Now loop over each of the simulated systems and fit to an harmonic system
        for i, (series, label) in enumerate(zip(serieslist, labels)):
            s = series[iseries]
            reffunc = -RT * np.log(s.density())
            p0 = [RT / s.var, s.mean, s.std * np.sqrt(np.pi * 2.0)]
            popt, pcov = opt.curve_fit(_harmonic, s.points, reffunc, p0)
            if isbond:
                print " %s: k=%9.3f r_eq=%9.3f" % (label, popt[0], popt[1]),
                ws.cell(row=rowoffset+3+iseries,column=4+i*3+1).value = popt[1] * 0.1
                ws.cell(row=rowoffset+3+iseries,column=4+i*3).value = popt[0] * 4.184 * 100
            else:
                print " %s: k=%9.3f theta_eq=%9.3f" % (label, popt[0], np.rad2deg(np.arccos(popt[1]))),
                ws.cell(row=rowoffset+3+iseries,column=4+i*3+1).value = np.rad2deg(np.arccos(popt[1]))
                ws.cell(row=rowoffset+3+iseries,column=4+i*3).value = popt[0] * 4.184
            ws.cell(row=rowoffset+3+iseries,column=4+i*3+2).value = s.shapiro()
        print ""


def _plot_dists(serieslist, labels, fflist, figure):
    """
    Plot the distribution of a series in subplots
    """
    nseries = len(fflist)
    ncols = 3.0
    nrows = np.ceil(nseries / ncols)
    RT = 303.0 * 8.314 / 4.184 / 1000.0

    for iseries in range(nseries):
        a = figure.add_subplot(nrows, ncols, iseries + 1)
        minval = 1000
        maxval = -1000
        stdval = 0
        # Plot the density for each simulated system
        lines = []
        for i, (series, label) in enumerate(zip(serieslist, labels)):
            s = series[iseries]
            x = s.points
            if isinstance(fflist[iseries].contype, elbalib.BondType):
                x = x * 0.1
            #l, = a.plot((x[1:]+x[:-1])/2.0, s.histo(), color=colors.color(i), label=label)
            l, = a.plot(x, s.density(), color=colors.color(i), label=label)
            lines.append(l)
            stdval = max(stdval, np.round(2.0 * s.std, 0), 0.5)
            minval = min(minval, np.floor(s.points.min()))
            maxval = max(maxval, np.ceil(s.points.max()))
        # Calculate the range of the x axis
        mean, std = fflist[iseries].contype.statmoments(RT)
        minval = min(minval, np.floor(mean - 2.0 * std))
        maxval = max(maxval, np.ceil(mean + 2.0 * std))
        # Plot the force field ideal distribution
        x = np.arange(minval, maxval, 0.1)
        y = fflist[iseries].contype.distribution(x, RT,trans=False)
        if isinstance(fflist[iseries].contype, elbalib.BondType):
            x = x * 0.1
        l, = a.plot(x, y, color=colors.color(len(labels)), label="ff")
        lines.append(l)
        stdval = max(stdval, np.round(2.0 * std, 0), 0.5)
        if isinstance(fflist[iseries].contype, elbalib.AngleType):
            minval = max(-1.0,minval)
            maxval = min(1.0,maxval)
            if stdval == 0.0 : stdval = np.cos(np.pi/4.0)

        # Add legend and ticks
        #if iseries == 0:
    #        a.legend(loc=3, fontsize=8,bbox_to_anchor = (0,1.02,1,0.102),
    #        mode="expand",ncol=2)
        a.set_yticks([])
        x = np.arange(minval, maxval+stdval, stdval)
        if isinstance(fflist[iseries].contype, elbalib.BondType):
            x = x * 0.1
        a.set_xticks(x)
        if isinstance(fflist[iseries].contype, elbalib.AngleType):
            a.set_xticklabels(np.rad2deg(np.arccos(x)))
            a.invert_xaxis()
        a.text(0.05,0.85,fflist[iseries].atomstring(),fontsize=8,transform=a.transAxes)
    labels2 = list(labels)
    labels2.append("ff")
    figure.legend(lines,labels2,loc=(0.03,0.95),fontsize=8,ncol=6)
    if isinstance(fflist[iseries].contype, elbalib.AngleType):
        figure.suptitle("Valance angle [deg]",y=0.05)
    else:
        figure.suptitle("Bond length [nm]",y=0.05)
    #figure.subplots_adjust(top=0.90)
    figure.tight_layout(rect=(0,0.025,1,0.96))


def _read_series(filename, nbonds):
    """
    Read series of data output by md_bondang.py
    """
    data = []
    with open(filename, "r") as f:
        line = f.readline()
        while line:
            data.append(line.strip().split())
            line = f.readline()
    data = np.array(data, dtype=float)
    bonddata = data[:, 1:nbonds + 1]
    angdata = data[:, nbonds + 1:]
    return bonddata, angdata


def _trans_ang(y,inv=False):
    if not inv:
        return np.cos(np.deg2rad(y))
    else:
        return np.rad2deg(np.arccos(y))

if __name__ == '__main__':

    print " ".join(sys.argv)

    parser = argparse.ArgumentParser(description="Analyse distribution of bonds and angles")
    parser.add_argument('-f', '--file', nargs="+", help="the trajectory of the bond and angles")
    parser.add_argument('-l', '--labels', nargs="+", help="the labels of the different trajectories")
    parser.add_argument('-r', '--ref', help="a CG reference file", default="ref.pdb")
    parser.add_argument('-x', '--xml', help="an XML file with force field definitions")
    parser.add_argument('-e','--excel',help="the filename of the XLSX file")
    parser.add_argument('-s','--sheet',help="the sheet in the XLSX file")
    parser.add_argument('-o', '--out', help="the output prefx", default="bondang")
    args = parser.parse_args()

    ff = elbalib.Elba()
    ff.load(args.xml)

    refuni = md.Universe(args.ref)

    # Find which residue the bond and angles were tracked
    ffmol = None
    for r in refuni.selectAtoms("all").residues:
        ffmol = ff.find_molecule(r.name)
        if ffmol is not None and ffmol.bonds and ffmol.angles:
            break
    if ffmol is None:
        raise Exception("Could not find any of the FF molecules in the reference structure")
    nbonds = len(ffmol.bonds)
    nangles = len(ffmol.angles)

    # Now we can read data, assume it was created by md_bondang.py
    anglist = []
    bondlist = []
    for filename in args.file:
        bd, ad = _read_series(filename, nbonds)
        # This transforms that numpy arrays into TimeSeries objects and thereby estimate densities etc
        bondlist.append([TimeSeries(bd[:, i]) for i in range(nbonds)])
        anglist.append([TimeSeries(ad[:, i], _trans_ang) for i in range(nangles)])

    # Create one plot for bonds and one for angles
    f = plt.figure(1)
    _plot_dists(bondlist, args.labels, ffmol.bonds, f)
    f.savefig(args.out + "_bonddist.png", format="png",dpi=300)

    f = plt.figure(2)
    _plot_dists(anglist, args.labels, ffmol.angles, f)
    f.savefig(args.out + "_angdist.png", format="png",dpi=300)

    # Estimate ff parameters
    try :
        wb = xl.load_workbook(filename = args.excel)
    except :
        print "Could not open the XLSX file. Will create one from scratch"
        wb = xl.Workbook()

    try :
        ws = wb[args.sheet]
    except :
        ws = wb.create_sheet(title=args.sheet)

    _fit_ff(bondlist, args.labels, ffmol.bonds, ws)
    print ""
    _fit_ff(anglist, args.labels, ffmol.angles, ws,offset=len(ffmol.bonds))
    wb.save(args.excel)
