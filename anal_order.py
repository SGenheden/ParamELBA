# Author: Samuel Genheden samuel.genheden@gmail.com

"""
Program to plot CG tail order parameters from CG and AA simulations
"""

import argparse
import os
import sys

import matplotlib.pylab as plt
import numpy as np
import openpyxl as xl
import scipy.stats as sstats

from sgenlib import colors

def _expand_selection(selection):
    """
    Expand a simple pattern with {A..B} format to an integer range
    """
    first = selection.index("{")
    last  = selection.index("}")
    start,end = map(int,selection[first+1:last].split(".."))
    return ["%s%d%s"%(selection[:first],i,selection[last+1:]) for i in range(start,end+1)]

def _plot_order(order_list,labels,figure):
    """
    Plot the distribution of a series in subplots
    """
    nchains = len(order_list[0])
    ncols = 2.0
    nrows = np.ceil(nchains / ncols)

    axes = []
    miny,maxy = 1.0,0.0
    for ichain in range(nchains):
        a = figure.add_subplot(nrows, ncols, ichain + 1)
        axes.append(a)
        for i,(simorder,label) in enumerate(zip(order_list,labels)):
            order = simorder[ichain]
            miny = min(miny,order.min())
            maxy = max(maxy,order.max())
            a.plot(np.arange(1,len(order)+1,1),order,'-*',label=label,color=colors.color(i))
        a.set_xticks(np.arange(0,len(order)+2,1))
        # Add legend and ticks
        if ichain == 0:
            a.legend(loc=1, fontsize=8)

    for a in axes :
        a.set_ylim([miny-0.05,maxy+0.05])

def _read_series(filename, chains_count):
    """
    Read series of data output by md_order.py
    """
    data = []
    with open(filename, "r") as f:
        line = f.readline()
        while line:
            data.append(line.strip().split())
            line = f.readline()
    data = np.array(data, dtype=float)
    # Split the data for the different chains
    chainorder = []
    for i,cnt in enumerate(chains_count) :
        first = 0 if i == 0 else chains_count[-1]-1
        last = first + cnt
        chainorder.append(data[:,first+1:last+1].mean(axis=0))
    return chainorder

if __name__ == '__main__':

    print " ".join(sys.argv)

    # Command-line input
    parser = argparse.ArgumentParser(description="Analyse CG order parameters")
    parser.add_argument('-f', '--file', nargs="+", help="the trajectory of the order parameters")
    parser.add_argument('-l', '--labels', nargs="+", help="the labels of the different trajectories")
    parser.add_argument('-c', '--chains',nargs="+", help="the carbon chains")
    parser.add_argument('-e','--excel',help="the filename of the XLSX file")
    parser.add_argument('-s','--sheet',help="the sheet in the XLSX file")
    parser.add_argument('-o', '--out', help="the output filename", default="order.png")
    args = parser.parse_args()

    # Parse the chain definition
    chains_beads = []
    chains_count = []
    for chain in args.chains :
        atoms = chain.split("-")
        if len(atoms) == 1: atoms = _expand_selection(chain)
        chains_beads.append(atoms)
        chains_count.append(len(chains_beads[-1])-1)

    # Read data, assume it was created by md_order.py
    order_list = []
    order_std_list = []
    for filename in args.file:
        chainorder = _read_series(filename, chains_count)
        h,t = os.path.split(filename)
        filename2 = os.path.join(h,"Repeat2",t)
        chainorder2 = _read_series(filename2,chains_count)
        chainorder_av = [np.asarray([d1,d2]).mean(axis=0)
                            for d1,d2 in zip(chainorder,chainorder2)]
        chainorder_std = [np.asarray([d1,d2]).std(axis=0)/np.sqrt(2)
                            for d1,d2 in zip(chainorder,chainorder2)]
        order_list.append(chainorder_av)
        order_std_list.append(chainorder_std)

    #
    f = plt.figure(1)
    _plot_order(order_list,args.labels,f)
    f.savefig(args.out,format="png")

    try :
        wb = xl.load_workbook(filename = args.excel)
    except :
        print "Could not open the XLSX file. Will create one from scratch"
        wb = xl.Workbook()

    try :
        ws = wb[args.sheet]
    except :
        ws = wb.create_sheet(title=args.sheet)
        
    for i,(label,order,orderstd) in enumerate(zip(args.labels,order_list,order_std_list),2):
        ws.cell(row=i,column=1).value = label
        for j,(o,ostd) in enumerate(zip(order,orderstd)):
            if i == 2:
                ws.cell(row=1,column=j*2+2).value = "chain%d"%(j+1)
            ws.cell(row=i,column=2+j*2).value = o.mean()
            ws.cell(row=i,column=2+j*2+1).value = np.sqrt(np.sum(ostd*ostd))
    wb.save(args.excel)
